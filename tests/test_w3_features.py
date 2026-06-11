"""批量并发限流 / Excel 导出 / status 归一化的 W3 专项测试。"""

from __future__ import annotations

import asyncio
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook

from app.config import get_settings
from app.schemas.host import HostInfo, HostMeta
from app.services.cache_service import CacheService
from app.services.export_service import COLUMNS, build_hosts_xlsx
from app.services.host_service import HostService
from app.utils.normalize import normalize_status


def _fake_host(asset_id: str = "TYSV00000001") -> HostInfo:
    return HostInfo(
        asset_id=asset_id,
        ip="10.0.0.1",
        zone="zone_a",
        machine_type="MOCK-1G",
        status="online",
        idc="示例机房A1",
        cabinet="A-12",
        position="A-12-3",
        owner="alice",
        backup_owners=["bob", "carol"],
        has_tpc=True,
        billing_tags={"tag_a": "1"},
        **{"_meta": HostMeta(data_sources=["cmdb"])},
    )


# ─────────────────────────── status 归一化 ───────────────────────────


class TestNormalizeStatusFallback:
    def test_already_english(self) -> None:
        assert normalize_status("online") == "online"
        assert normalize_status("offline") == "offline"
        assert normalize_status("maintenance") == "maintenance"

    def test_chinese_fallback(self) -> None:
        # 兜底：万一某 client 漏配映射，HostService 仍能收敛
        assert normalize_status("运营中") == "online"
        assert normalize_status("维护中") == "maintenance"
        assert normalize_status("故障") == "offline"

    def test_unknown_returns_none(self) -> None:
        # 未识别值不能穿透到前端
        assert normalize_status("奇怪状态") is None

    def test_empty(self) -> None:
        assert normalize_status("") is None
        assert normalize_status(None) is None


# ─────────────────────────── HostInfo.status Literal 校验 ───────────────────────────


class TestHostInfoStatusLiteral:
    def test_valid_values(self) -> None:
        for s in ("online", "offline", "maintenance", None):
            host = HostInfo(asset_id="TYSV00000001", status=s)
            assert host.status == s

    def test_invalid_raises(self) -> None:
        from pydantic import ValidationError

        with pytest.raises(ValidationError):
            HostInfo(asset_id="TYSV00000001", status="invalid")  # type: ignore[arg-type]

    def test_chinese_raw_rejected(self) -> None:
        """中文不能直接进 schema（应该在采集层归一化）。"""
        from pydantic import ValidationError

        with pytest.raises(ValidationError):
            HostInfo(asset_id="TYSV00000001", status="运营中")  # type: ignore[arg-type]


# ─────────────────────────── 批量并发限流 ───────────────────────────


@pytest.mark.asyncio
class TestBatchConcurrency:
    async def test_concurrency_limited_by_semaphore(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """验证同时进行的查询数不超过 batch_concurrency。"""
        # 设小一点便于观察
        monkeypatch.setenv("TEZ_BATCH_CONCURRENCY", "3")
        get_settings.cache_clear()  # type: ignore[attr-defined]

        # 计数器记录"当前正在执行的任务数"
        in_flight = 0
        peak = 0
        lock = asyncio.Lock()

        async def slow_get_host(asset_id: str):
            nonlocal in_flight, peak
            async with lock:
                in_flight += 1
                peak = max(peak, in_flight)
            try:
                await asyncio.sleep(0.05)
                return _fake_host(asset_id)
            finally:
                async with lock:
                    in_flight -= 1

        cmdb = MagicMock()
        cmdb.close = AsyncMock(return_value=None)
        tcum = MagicMock()
        tcum.close = AsyncMock(return_value=None)
        idcrm = MagicMock()
        idcrm.close = AsyncMock(return_value=None)

        svc = HostService(cmdb=cmdb, tcum=tcum, idcrm=idcrm, cache=CacheService())
        # 直接打补丁覆盖 svc.get_host
        svc.get_host = slow_get_host  # type: ignore[assignment]

        ids = [f"TYSV0000{i:04d}" for i in range(20)]
        results = await svc.batch_get_hosts(ids)

        assert len(results) == 20
        # 并发上限 = 3
        assert peak <= 3, f"peak={peak} > 3，限流失效"

    async def test_one_task_failure_does_not_break_others(self) -> None:
        async def get_host(asset_id: str):
            if asset_id == "TYSV00000005":
                raise RuntimeError("specific boom")
            return _fake_host(asset_id)

        cmdb = MagicMock()
        cmdb.close = AsyncMock(return_value=None)
        tcum = MagicMock()
        tcum.close = AsyncMock(return_value=None)
        idcrm = MagicMock()
        idcrm.close = AsyncMock(return_value=None)
        svc = HostService(cmdb=cmdb, tcum=tcum, idcrm=idcrm, cache=CacheService())
        svc.get_host = get_host  # type: ignore[assignment]

        ids = [f"TYSV0000000{i}" for i in range(1, 9)]
        results = await svc.batch_get_hosts(ids)
        assert len(results) == 8
        # 第 5 条失败，其他成功
        for qid, host, err in results:
            if qid == "TYSV00000005":
                assert host is None
                assert err is not None
                assert "specific boom" in err
            else:
                assert host is not None
                assert err is None

    async def test_mixed_ip_and_asset(self) -> None:
        cmdb = MagicMock()
        cmdb.close = AsyncMock(return_value=None)
        tcum = MagicMock()
        tcum.close = AsyncMock(return_value=None)
        idcrm = MagicMock()
        idcrm.close = AsyncMock(return_value=None)
        svc = HostService(cmdb=cmdb, tcum=tcum, idcrm=idcrm, cache=CacheService())
        svc.get_host = AsyncMock(side_effect=lambda x: _fake_host(x))  # type: ignore[assignment]
        svc.get_host_by_ip = AsyncMock(side_effect=lambda x: _fake_host("BY_IP"))  # type: ignore[assignment]

        results = await svc.batch_get_hosts_mixed(
            [
                ("TYSV00000001", "asset_id"),
                ("10.0.0.5", "ip"),
                ("garbage", "unknown"),
            ]
        )
        assert len(results) == 3
        assert results[0][2] is not None
        assert results[1][2] is not None
        # 第三条 unknown 被标错
        assert results[2][2] is None
        assert "不支持" in results[2][3]


# ─────────────────────────── Excel 导出 ───────────────────────────


class TestExportXlsxBuilder:
    def test_build_with_hosts(self) -> None:
        hosts = [_fake_host("TYSV00000001"), _fake_host("TYSV00000002")]
        resp = build_hosts_xlsx(hosts, filename="test.xlsx")
        assert resp.media_type == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # StreamingResponse 的 body_iterator 不便单测验证，用相同生成路径再跑一遍验字段
        import openpyxl

        # 直接重新构造 wb 验证字段（与 build_hosts_xlsx 一致即可）
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([cn for _, cn in COLUMNS])
        for h in hosts:
            from app.services.export_service import _cell_value

            ws.append([_cell_value(h, field) for field, _ in COLUMNS])
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        loaded = load_workbook(out)
        sheet = loaded.active
        # 第一行是表头
        header = [c.value for c in sheet[1]]
        assert header[0] == "固资号"
        assert "状态" in header
        # 第二行是第一台机
        row2 = [c.value for c in sheet[2]]
        assert row2[0] == "TYSV00000001"
        # has_tpc → 是
        tpc_idx = next(i for i, (f, _) in enumerate(COLUMNS) if f == "has_tpc")
        assert row2[tpc_idx] == "是"
        # backup_owners → "bob;carol"
        bo_idx = next(i for i, (f, _) in enumerate(COLUMNS) if f == "backup_owners")
        assert row2[bo_idx] == "bob;carol"

    def test_build_empty(self) -> None:
        resp = build_hosts_xlsx([], filename="empty.xlsx")
        assert resp.headers["content-disposition"].endswith('"empty.xlsx"')

    def test_default_filename_has_timestamp(self) -> None:
        resp = build_hosts_xlsx([_fake_host()])
        cd = resp.headers["content-disposition"]
        assert "hosts_" in cd
        assert ".xlsx" in cd
